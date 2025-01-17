from openpyxl import load_workbook
from openpyxl import Workbook

# Load the workbook
file_path = "gdp_modified.xlsx"
workbook = load_workbook(filename=file_path)

# Access the active sheet
sheet = workbook.active

# Delete the first 8 rows
sheet.delete_rows(1, 8)

# Save the modified workbook
output_file_path = "gdp_modified.xlsx"
workbook.save(filename=output_file_path)

# Read the remaining data
data = []
for row in sheet.iter_rows(values_only=True):
    data.append(row)

# Process the data to remove "Q" from year values and filter out years before 1980
processed_data = []
for row in data:
    year = row[0]
    if isinstance(year, str) and "Q" in year:
        year = year.split("Q")[0].strip()  # Remove "Q" and any trailing spaces
    # Convert year to integer for comparison
    year_int = int(year)
    # Only include rows where the year is 1980 or later
    if year_int >= 1980:
        processed_data.append((year, row[1]))

# Remove the duplicate rows appearing at the top
processed_data = processed_data[12:]  # Skip the first 12 rows

# Create a new workbook for cleaned data
new_workbook = Workbook()
new_sheet = new_workbook.active

# Write the processed data to the new sheet
for row in processed_data:
    new_sheet.append(row)

# Save the cleaned data to a new file
new_file_path = "cleaned_gdp.xlsx"
new_workbook.save(filename=new_file_path)

# Print all the cleaned data
print("Cleaned Data:")
for row in processed_data:
    print(row)

# Close the workbooks
workbook.close()
new_workbook.close()

print(f"Cleaned data saved as {new_file_path}")
